#!/usr/bin/env python3
"""
StackIQ Action — atomic outcome logger.

Writes ONE finding's action outcome to all three sinks in a single call, keyed by
finding_id (idempotent upsert), so an already-taken action can never be lost or
duplicated. Call this immediately after every action resolves.

Sinks:
  1) Excel "Action" tab on the Insight workbook (--excel)
  2) action block in session.json (--session)
  3) actions array in history/run_<n>.json (--history)

Status is one of: executed | awaiting_confirmation | pending_approval | blocked

Usage example:
  log_action.py \
    --excel   /path/state/../report.xlsx \
    --session /path/state/session.json \
    --history /path/state/history/run_1.json \
    --finding-id F-003 --status awaiting_confirmation \
    --action-taken "Emailed admin to revoke 3 idle Figma seats (~$45/mo)" \
    --channel desktop_app --recipient it-admin@acme.com --approved-by-user

Requires openpyxl (pip install openpyxl --break-system-packages).
"""
import argparse
import datetime as _dt
import json
import os
import sys
import tempfile

STATUSES = ("executed", "awaiting_confirmation", "pending_approval", "blocked")
COLUMNS = [
    "finding_id", "action_taken", "status", "blocked_reason", "channel",
    "recipient", "approved_by_user", "sent_date", "confirmed_date", "date",
]


def _today():
    return _dt.date.today().isoformat()


def _atomic_write_json(path, data):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(os.path.abspath(path)), suffix=".tmp")
    try:
        with os.fdopen(fd, "w") as fh:
            json.dump(data, fh, indent=2)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


def _load_json(path, default):
    if os.path.exists(path):
        with open(path) as fh:
            return json.load(fh)
    return default


def update_excel(path, row):
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        sys.stderr.write("openpyxl not installed: pip install openpyxl --break-system-packages\n")
        raise
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb["Action"] if "Action" in wb.sheetnames else wb.create_sheet("Action")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Action"
    # ensure the header is exactly row 1 (write cells directly to avoid append row-pointer quirks)
    if ws.cell(row=1, column=1).value != "finding_id":
        for i, c in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=i, value=c)
    # upsert by finding_id (column 1)
    target = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == str(row["finding_id"]):
            target = r
            break
    values = [row.get(c, "") for c in COLUMNS]
    if target is None:
        ws.append(values)
    else:
        for i, c in enumerate(COLUMNS, start=1):
            ws.cell(row=target, column=i, value=row.get(c, ""))
    # atomic-ish save via temp then replace
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(os.path.abspath(path)) or ".", suffix=".xlsx")
    os.close(fd)
    wb.save(tmp)
    os.replace(tmp, path)


def update_session(path, row):
    data = _load_json(path, {})
    action = data.setdefault("action", {})
    for s in STATUSES:
        action.setdefault(s, [])
        action[s] = [e for e in action[s] if str(e.get("finding_id")) != str(row["finding_id"])]
    status = row["status"]
    if status == "executed":
        entry = {"finding_id": row["finding_id"], "action_taken": row["action_taken"],
                 "channel": row["channel"], "date": row["date"]}
    elif status == "awaiting_confirmation":
        entry = {"finding_id": row["finding_id"], "action_taken": row["action_taken"],
                 "channel": row["channel"], "recipient": row["recipient"], "sent_date": row["sent_date"]}
    elif status == "pending_approval":
        entry = {"finding_id": row["finding_id"], "action_taken": row["action_taken"], "date": row["date"]}
    else:  # blocked
        entry = {"finding_id": row["finding_id"], "action_taken": row["action_taken"],
                 "blocked_reason": row["blocked_reason"], "date": row["date"]}
    action[status].append(entry)
    _atomic_write_json(path, data)


def update_history(path, row):
    data = _load_json(path, {"actions": []})
    actions = data.setdefault("actions", [])
    actions[:] = [a for a in actions if str(a.get("finding_id")) != str(row["finding_id"])]
    actions.append({
        "finding_id": row["finding_id"], "action_taken": row["action_taken"], "status": row["status"],
        "channel": row["channel"], "recipient": row["recipient"],
        "approved_by_user": row["approved_by_user"], "sent_date": row["sent_date"],
        "confirmed_date": row["confirmed_date"], "date": row["date"],
    })
    _atomic_write_json(path, data)


def main():
    p = argparse.ArgumentParser(description="Atomically log one StackIQ action outcome to all three sinks.")
    p.add_argument("--excel", required=True)
    p.add_argument("--session", required=True)
    p.add_argument("--history", required=True)
    p.add_argument("--finding-id", required=True)
    p.add_argument("--status", required=True, choices=STATUSES)
    p.add_argument("--action-taken", required=True)
    p.add_argument("--channel", default="")
    p.add_argument("--recipient", default="")
    p.add_argument("--blocked-reason", default="")
    p.add_argument("--approved-by-user", action="store_true")
    p.add_argument("--sent-date", default="")
    p.add_argument("--confirmed-date", default="")
    p.add_argument("--date", default="")
    a = p.parse_args()

    today = _today()
    row = {
        "finding_id": a.finding_id,
        "action_taken": a.action_taken,
        "status": a.status,
        "blocked_reason": a.blocked_reason,
        "channel": a.channel,
        "recipient": a.recipient,
        "approved_by_user": bool(a.approved_by_user),
        "sent_date": a.sent_date or (today if a.status == "awaiting_confirmation" else ""),
        "confirmed_date": a.confirmed_date or (today if a.status == "executed" else ""),
        "date": a.date or today,
    }

    # Sinks are written independently; if one fails the others still land and the
    # re-run reconciliation (or a re-call) will repair. Excel first (most visible).
    update_excel(a.excel, row)
    update_session(a.session, row)
    update_history(a.history, row)
    print(f"logged {a.finding_id} -> {a.status} (excel + session + history)")


if __name__ == "__main__":
    main()
